# 抓取url = 'https://github.com/apache?page=1'页面所有30个项目的第一页commits相关信息
# 抓取提交时间时发现部分提交时间内容有一个很奇怪的问题，如下：
# 网页上信息为<relative-time datetime="2018-11-15T19:40:22Z">Nov 16, 2018</relative-time>
#   抓取到的为<relative-time datetime="2018-11-15T19:40:22Z">Nov 15, 2018</relative-time>
# 猜测是时区不同造成的。(*^__^*)

from bs4 import BeautifulSoup
import requests
import re
from openpyxl import Workbook
import time

wb = Workbook()
ws = wb.active

def get_pro_rul(url):
    r = requests.get(url).text
    # 编写正则表达式从html文件中找到项目网址后半部分
    r_pattern = "\"\/apache\/\w+\">|\"\/apache\/\w+\-\w+\">|\"\/apache\/\w+\-\w+\-\w+\">|\"\/apache\/\w+\-\w+\-\w+\-\w+\">|\"\/apache\/\w+\-\w+\-\w+\-\w+\-\w+\">|\"\/apache\/\w+\-\w+\-\w+\-\w+\-\w+\-\w+\">"
    temp = re.findall(r_pattern, r)

    # 对所得结果处理，得到项目url后半部分
    find_pro_url = []
    for i in range(len(temp)):
        tem1 = temp[i]
        tem2 = tem1[1:-2]
        find_pro_url.append(tem2)

    # 项目url拼接得到每个项目commits的url
    pro_url = []
    for i in range(0, len(find_pro_url)):
        pro_url.append('https://github.com' + find_pro_url[i])
    return (pro_url)


def pro_commits_url_and_result(pro_url, length):
    # 获取每个项目commits页url
    # 设置循环计数器，控制表格写入
    count = 1
    pro_commits_url = []
    for i in range(length):
        url = pro_url[i]
        # 得到commit大页面
        pro_name = url[26:]
        # print(f"firt print {pro_name}")
        r = requests.get(url).text
        soup = BeautifulSoup(r, 'html.parser')
        # 获取该网页中href符合指定指定正则表达式的内容
        href_ = soup.find_all(href=re.compile(
            f"\/apache\/{pro_name}\/commits"))
        for each in href_:
            if str(each.get('href'))[:4] == 'http':
                # 处理得到commits url
                temp_url = each.get('href')
                pro_commits_url.append(temp_url[:-5])

        # 在项目的commits页面寻找内涵的第一页所有commits的url
        url = pro_commits_url[i]
        r = requests.get(url).text
        soup = BeautifulSoup(r, 'html.parser')

        myAttrs = {'class': 'message js-navigation-open'}
        find = soup.find_all(name='a', attrs=myAttrs)
        url_before = 'https://github.com'
        commits_user_url = []
        for each_temp in find:
            if str(each.get('href'))[:7] == '/apache':
                # 处理得到commits url
                temp_url = each_temp.get('href')
                temp_url = url_before + temp_url
                commits_user_url.append(temp_url)

        # 去掉所有重复获得的url（可能出现多个人共同提交一个commits，导致多个相同的commits网址被找到）
        commits_user_url_used=[]
        for each_url in commits_user_url:
            if each_url not in commits_user_url_used:
                commits_user_url_used.append(each_url)
        # commits_user_url_used = list(set(commits_user_url))

        # 对具体提交commits页面进行抓取数据
        for j in range(0, len(commits_user_url_used)):
            url = commits_user_url_used[j]
            r = requests.get(url).text
            soup = BeautifulSoup(r, 'html.parser')

            # 抓取提交人
            # 提交人内容可能在不同标签内，处理并抓取
            myAttrs = {'class': 'commit-author tooltipped tooltipped-s user-mention'}
            find_1 = soup.find_all(name='a', attrs=myAttrs)
            myAttrs = {'class': 'commit-author user-mention'}
            find_2 = soup.find_all(name='span', attrs=myAttrs)
            html_doc_author = f"{find_1}{find_2}"
            soup_new_author = BeautifulSoup(html_doc_author, 'html.parser')
            author_name_1 = []
            if len(find_1) != 0:
                for k in soup_new_author.find_all('a'):
                    author_name_1.append(k.string)
            author_name_2 = []
            if len(find_2) != 0:
                for k in soup_new_author.find_all('span'):
                    author_name_2.append(k.string)
            # 可能会有多个人一起提交一个commits，对此作出处理
            author_name_list = []
            if len(author_name_1) != 0:
                for each in author_name_1:
                    author_name_list.append(each)
            if len(author_name_2) != 0:
                for each in author_name_2:
                    author_name_list.append(each)
            # 装载提交人名字的list转化成string方便写入excel
            if len(author_name_list) < 2:
                author_name = author_name_list[0]
            elif len(author_name_list) < 3:
                    author_name = author_name_list[0] + " 、" + author_name_list[1]
            else:
                print("提交人超过两个，需要修改程序。")

            # 抓取其他需要信息
            relative_time_tag = soup.find_all('relative-time')
            myAttrs = {'class': 'toc-diff-stats'}
            file_change_tag = soup.find_all(name='div', attrs=myAttrs)
            myAttrs = {'class': 'sha user-select-contain'}
            item_number_tag = soup.find_all(name='span', attrs=myAttrs)
            html_doc = f"{relative_time_tag}{file_change_tag}{item_number_tag}"
            soup_new = BeautifulSoup(html_doc, 'html.parser')
            for k in soup_new.find_all('relative-time'):
                relative_time = k.string
            for k in soup_new.find_all('span'):
                item_number = k.string
            # 抓取文件数并处理
            for k in soup_new.find_all('button'):
                file_change = k.string
                file_change = [int(s)
                               for s in file_change.split() if s.isdigit()]
            # 增加代码行和删除代码行一起放入一个list
            file = []
            for k in soup_new.find_all('strong'):
                file.append(k.string)
            # 提取增加代码行和删除代码数目
            add = file[0].translate(str.maketrans('', '', 'additions'))
            dellet = file[1].translate(str.maketrans('', '', 'deletions'))
            title = ['项目名称', '提交人', '提交时间', '提交编号', '文件数', '增加代码行', '删除代码行']
            data = [str(pro_name), author_name, relative_time,
                    item_number, file_change[0], add, dellet]
            # 写入xlsx表格
            for k in range(1, 8):
                ws.cell(column=k, row=1).value = title[k - 1]
                ws.cell(column=k, row=count + 1).value = data[k - 1]


            count = count + 1
            print(count)
            print(time.process_time())
            wb.save('GitHub_get.xlsx')
    wb.save('GitHub_get.xlsx')


# 执行程序，获得结果
if __name__ == '__main__':
    url = 'https://github.com/apache?page=1'
    pro_url = get_pro_rul(url)
    length = len(pro_url)
    pro_commits_url_and_result(pro_url, length)
