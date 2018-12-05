from bs4 import BeautifulSoup
import requests
import re

import time
from openpyxl import Workbook


wb = Workbook()
ws = wb.active

# 得到一个项目commits中第一页所属的所有提交项目url
url = 'https://github.com/apache/ignite/commits/master'
r = requests.get(url).text
soup = BeautifulSoup(r, 'html.parser')

myAttrs = {'class': 'message js-navigation-open'}
find = soup.find_all(name='a', attrs=myAttrs)
url_before = 'https://github.com'
commits_user_url = []
for each in find:
    if str(each.get('href'))[:7] == '/apache':
        # 处理得到commits url
        temp_url = each.get('href')
        temp_url = url_before + temp_url
        commits_user_url.append(temp_url)

print(commits_user_url)
print(len(commits_user_url))
url_2=[]
for each_url in commits_user_url:
    if each_url not in url_2:
        url_2.append(each_url)
# url_2 = list(set(commits_user_url))
print(url_2)
print(len(url_2))


for i in range(0, len(url_2)):
    url = url_2[i]
    r = requests.get(url).text
    soup = BeautifulSoup(r, 'html.parser')


    # author
    myAttrs = {'class': 'commit-author tooltipped tooltipped-s user-mention'}
    find_1 = soup.find_all(name='a', attrs=myAttrs)
    myAttrs = {'class': 'commit-author user-mention'}
    find_2 = soup.find_all(name='span', attrs=myAttrs)

    html_doc_author = f"{find_1}{find_2}"
    soup_new_author = BeautifulSoup(html_doc_author, 'html.parser')
    author_name_1 = []
    if len(find_1) != 0:
        for k in soup_new_author.find_all('a'):
            author_name_1.append(k.string)

    author_name_2 = []
    if len(find_2) != 0:
        for k in soup_new_author.find_all('span'):
            author_name_2.append(k.string)

    author_name = []
    if len(author_name_1) != 0:
        for each in author_name_1:
            author_name.append(each)
    if len(author_name_2) != 0:
        for each in author_name_2:
            author_name.append(each)
    print(author_name)

    if len(author_name) > 1:
        author = author_name[0] + "、" + author_name[1]
    else:
        author = author_name[0]




    # upload time
    up_time = soup.find_all('relative-time')

    # 更改文件数目
    myAttrs = {'class': 'toc-diff-stats'}
    file_change_tag = soup.find_all(name='div', attrs=myAttrs)

    # 提交号
    myAttrs = {'class': 'sha user-select-contain'}
    item_number = soup.find_all(name='span', attrs=myAttrs)

    #
    html_doc = f"{up_time}{file_change_tag}{item_number}"
    soup_new = BeautifulSoup(html_doc, 'html.parser')

    for k in soup_new.find_all('relative-time'):
        up_time = k.string
    for k in soup_new.find_all('span'):
        item_number = k.string
    for k in soup_new.find_all('button'):
        file_change = k.string
        file_change = [int(s) for s in file_change.split() if s.isdigit()]
    file = []
    for k in soup_new.find_all('strong'):
        file.append(k.string)

    pro_name = "ignite"

    title = ['项目名称', '提交人', '提交时间', '提交编号', '文件数', '增加代码行', '删除代码行']
    add = file[0].translate(str.maketrans('', '', 'additions'))
    dellet = file[1].translate(str.maketrans('', '', 'deletions'))
    data = [str(pro_name), str(author), str(up_time), str(
        item_number), file_change[0], add, dellet]
    for k in range(1, 8):
        ws.cell(column=k, row=1).value = title[k - 1]
        ws.cell(column=k, row=i + 2).value = data[k - 1]

    # ws.cell(column=1, row=1).value = '项目名称'
    # ws.cell(column=2, row=1).value = '提交人'
    # ws.cell(column=3, row=1).value = '提交时间'
    # ws.cell(column=4, row=1).value = '提交编号'
    # ws.cell(column=5, row=1).value = '文件数'
    # ws.cell(column=6, row=1).value = '增加代码行'
    # ws.cell(column=7, row=1).value = '删除代码行'
    # ws.cell(column=1, row=i + 1).value = str(pro_name)
    # ws.cell(column=2, row=i + 1).value = str(author)
    # ws.cell(column=3, row=i + 1).value = str(up_time)
    # ws.cell(column=4, row=i + 1).value = str(item_number)
    # ws.cell(column=5, row=i + 1).value = file_change[0]
    # ws.cell(column=6, row=i + 1).value = file[0].translate(str.maketrans('','','additions'))
    # ws.cell(column=7, row=i + 1).value = file[1].translate(str.maketrans('','','deletions'))
    print(i)
    print(time.process_time())

wb.save('2.xlsx')


print(time.process_time())
