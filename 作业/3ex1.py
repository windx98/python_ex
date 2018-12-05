# 抓取url = 'https://github.com/apache?page=1'页面所有30个项目的第一页commits相关信息
# 抓取提交时间时发现部分提交时间内容有一个很奇怪的问题，如下：
# 网页上信息为<relative-time datetime="2018-11-15T19:40:22Z">Nov 16, 2018</relative-time>
#   抓取到的为<relative-time datetime="2018-11-15T19:40:22Z">Nov 15, 2018</relative-time>

from bs4 import BeautifulSoup
import requests
import re
from openpyxl import Workbook
import string

wb = Workbook()
ws = wb.active

# 获得项目地址
def get_pro_rul(url):
    r = requests.get(url).text
    soup_commits = BeautifulSoup(r, 'html.parser')

    url_head ='https://github.com'
    # 首先抓取符合如下特征的标签
    myAttrs = {'class': 'd-inline-block mb-1'}
    item_number = soup_commits.find_all(name='div', attrs=myAttrs)
    # 对抓取到的标签进一步处理得到装有url
    pro_url = []
    for each in item_number:
        url = each.find_all('a')
        pro_url.append(url_head + url[0].get('href'))
    return (pro_url)

# 得到有需要信息的页面url并抓取信息
def pro_commits_url_and_result(pro_url, length):
    # 获取每个项目commits页url
    # 设置循环计数器，控制表格写入
    count = 1
    pro_commits_url = []
    for i in range(length):
        url = pro_url[i]
        # 得到commit大页面url
        pro_name = url[26:]
        r = requests.get(url).text
        soup = BeautifulSoup(r, 'html.parser')
        # 获取该网页中href符合指定指定正则表达式的内容
        href_ = soup.find_all(href=re.compile(
            f"\/apache\/{pro_name}\/commits"))
        for each in href_:
            if str(each.get('href'))[:4] == 'http':
                # 处理得到commits url
                temp_url = each.get('href')
                pro_commits_url.append(temp_url[:-5])

        # 在项目的commits页面寻找内涵的第一页所有commits的url
        url = pro_commits_url[i]
        r = requests.get(url).text
        soup = BeautifulSoup(r, 'html.parser')

        myAttrs = {'class': 'message js-navigation-open'}
        find = soup.find_all(name='a', attrs=myAttrs)
        url_before = 'https://github.com'
        commits_user_url = []
        for each_temp in find:
            if str(each.get('href'))[:7] == '/apache':
                # 处理得到commits url
                temp_url = each_temp.get('href')
                temp_url = url_before + temp_url
                commits_user_url.append(temp_url)

        # 去掉所有重复获得的url（可能出现多个人共同提交一个commits，导致多个相同的commits网址被找到）
        commits_user_url_used=[]
        for each_url in commits_user_url:
            if each_url not in commits_user_url_used:
                commits_user_url_used.append(each_url)

        # 对具体提交commits页面进行抓取数据
        for j in range(0, len(commits_user_url_used)):
            url = commits_user_url_used[j]
            r = requests.get(url).text
            soup = BeautifulSoup(r, 'html.parser')

            # 提交人内容可能在不同标签内，处理并抓取（标签名和提交者设置和人数有关）
            myAttrs = {'class': 'commit-author tooltipped tooltipped-s user-mention'}
            find_1 = soup.find_all(name='a', attrs=myAttrs)
            myAttrs = {'class': 'commit-author user-mention'}
            find_2 = soup.find_all(name='span', attrs=myAttrs)
            html_doc_author = f"{find_1}{find_2}"
            soup_new_author = BeautifulSoup(html_doc_author, 'html.parser')
            author_name_1 = []
            if len(find_1) != 0:
                for k in soup_new_author.find_all('a'):
                    author_name_1.append(k.string)
            author_name_2 = []
            if len(find_2) != 0:
                for k in soup_new_author.find_all('span'):
                    author_name_2.append(k.string)
            # 可能会有多个人一起提交一个commits，对此作出处理
            author_name_list = []
            if len(author_name_1) != 0:
                for each in author_name_1:
                    author_name_list.append(each)
            if len(author_name_2) != 0:
                for each in author_name_2:
                    author_name_list.append(each)
            # 解决三个人一起提交，暂时还没看到4个
            if len(author_name_list) == 0:
                myAttrs = {'class': 'text-bold'}
                find_3 = soup.find_all(name='span', attrs=myAttrs)
                text = find_3[0].get('title')
            # 装载提交人名字的list转化成string方便写入excel
            if len(author_name_list) != 0:
                if len(author_name_list) < 2:
                    author_name = author_name_list[0]
                elif len(author_name_list) < 3:
                    author_name = author_name_list[0] + " 、" + author_name_list[1]
            else:
                author_name = text

            # 抓取其他需要信息
            relative_time_tag = soup.find_all('relative-time')

            myAttrs = {'class': 'toc-diff-stats'}
            file_change_tag = soup.find_all(name='div', attrs=myAttrs)
            myAttrs = {'class': 'sha user-select-contain'}
            item_number_tag = soup.find_all(name='span', attrs=myAttrs)
            html_doc = f"{relative_time_tag}{file_change_tag}{item_number_tag}"
            soup_new = BeautifulSoup(html_doc, 'html.parser')
            for k in soup_new.find_all('relative-time'):
                relative_time = k.string
            for k in soup_new.find_all('span'):
                item_number = k.string
            # 提取文件数目，代码增加和删除数目。
            # 因为一些未知原因，部分页面存放文件数目的标签不同于大多数情况，加以修正。
            file = []
            for k in soup_new.find_all('strong'):
                file.append(k.string)
            if len(file) != 3:
                for k in soup_new.find_all('button'):
                    file_change = k.string
                    file_change = [int(s)for s in file_change.split() if s.isdigit()]
                    file_change = file_change[0]
                    add = file[0].split()[0]
                    dellet = file[1].split()[0]
            else:
                file_change = file[0].split()[0]
                print(file_change)
                add = file[1].split()[0]
                dellet = file[2].split()[0]

            title = ['项目名称', '提交人', '提交时间', '提交编号', '文件数', '增加代码行', '删除代码行']
            data = [str(pro_name), author_name, relative_time,
                    item_number, str(file_change), add, dellet]
            # 写入xlsx表格
            for k in range(1, 8):
                ws.cell(column=k, row=1).value = title[k - 1]
                ws.cell(column=k, row=count + 1).value = data[k - 1]
            count = count + 1
            print(count)
            # 随时保存。
            wb.save('GitHub.xlsx')
    wb.save('GitHub.xlsx')


# 执行程序，获得结果
if __name__ == '__main__':
    url = 'https://github.com/apache?page=1'
    pro_url = get_pro_rul(url)
    length = len(pro_url)
    pro_commits_url_and_result(pro_url, length)

