from openpyxl import Workbook
from openpyxl.styles import Font, colors, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string

wb = Workbook()
ws = wb.create_sheet("my", 0)

# 设置指定区域内单元格上下左右居中并修改行高列宽为指定值
for row in range(1, 9):
    for col in range(1, 9):
        # 单元格内内容居中
        ws.cell(column=col, row=row).alignment = Alignment(
            horizontal='center', vertical='center')
        # 设置行高列宽为指定值
        ws.row_dimensions[row].height = 35
        # 将数字转化为字母以供下面调用
        c_char = get_column_letter(col)
        ws.column_dimensions[c_char].width = 20

# 写入内容
ws.merge_cells('A1:H1')
ws["A1"] = '时间安排表'
ws.merge_cells('A2:H2')
ws["A2"] = '时间：2018年下'
ws['B3'] = '星期一'
ws['C3'] = '星期二'
ws['D3'] = '星期三'
ws['E3'] = '星期四'
ws['F3'] = '星期五'
ws['G3'] = '星期六'
ws['H3'] = '星期天'

ws['A4'] = '1，2节'
ws['B4'] = '材料热力学'
ws['C4'] = '材料测试分析方法'
ws['D4'] = '材料热力学'
ws['E4'] = '材料测试分析方法'
ws['F4'] = '复合材料原理'

ws.merge_cells('G4:G8')
ws["G4"] = '周末休息'
ws.merge_cells('H4:H8')
ws["H4"] = '周末休息'

ws['A5'] = '3，4节'
ws['B5'] = '复习'
ws['C5'] = '界面物理化学'
ws['D5'] = '机械设计基础'
ws['E5'] = '界面物理化学'
ws['F5'] = '检测技术与控制工程'

ws['A6'] = '5，6节'
ws['B6'] = '机械设计基础'
ws['C6'] = '材料科学研究方法'
ws['F6'] = '微机原理'
ws.merge_cells('D6:D7')
ws['D6'] = '复习'
ws.merge_cells('E6:E8')
ws['E6'] = '复习'


ws['A7'] = '7，8节'
ws['B7'] = '微机原理'
ws.merge_cells('C7:C8')
ws['C7'] = '阅读'
ws.merge_cells('F7:F8')
ws['F7'] = '阅读'

ws['A8'] = '9，10节'
ws['B8'] = 'python编程与科学计算'
ws['D8'] = '国际商务礼仪'


wb.save('日程表.xlsx')
